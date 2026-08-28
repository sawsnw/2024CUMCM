"""问题四求解脚本：S 形调头曲线优化与板凳龙运动计算。

运行方式:
    python scripts/solve_problem4.py

功能:
    1. 求解“基准（给定）”配置的 S 形调头曲线几何（两圆弧与调头空间圆内切），
       得到基准调头曲线长度 L0；
    2. 以两个切点参数 (θ1, θ2) 为决策变量做二维几何粗搜索 + 多起点局部细化，
       并按弧长从小到大对候选进行全时段碰撞检验，找到最短无碰撞 S 形曲线 L*，
       得到缩短量 ΔL 与缩短比例 η（文档第 7、8 节）；
    3. 对最优配置建立弧长参数化完整路径 Γ(s)，递推 t=-100..100 s 全部 224 个
       把手的位置与速度（文档第 9~11 节）；
    4. 读取模板 data/result4.xlsx，将结果写入 results/tables/result4.xlsx
       （不修改模板，覆盖旧输出，保留 6 位小数）；
    5. 运行模型检验（相切/连续性、半径比、区域约束、刚性距离、速度约束、碰撞）；
    6. 输出基准/最优调头曲线对比图 results/figures/problem4_turning_comparison.png；
    7. 日志写入 results/logs/solve_problem4_YYYY-MM-DD_HHMMSS.log（不覆盖旧文件）。
"""

import concurrent.futures
import os
import time

import numpy as np
import openpyxl

from src.config import (
    P4_RESULT_PATH,
    P4_RESULT_TEMPLATE_PATH,
    TOTAL_HANDLES,
    TURN_RADIUS,
)
from src.models.collision import global_margin
from src.models.problem4 import (
    ORIENTATION_MINUS,
    ORIENTATION_PLUS,
    compute_dragon_at,
    optimize_turning_curve,
    path_point,
    path_tangent,
    s_curve_points,
    solve_given_configuration,
    solve_s_curve,
    turning_candidate,
)
from src.models.spiral_dragon import handle_distances
from src.utils.logger import setup_logger
from src.visualization.problem4_plots import plot_turning_comparison

# 时间范围与步长
T_START, T_END = -100.0, 100.0
TIME_STEP = 1.0
# 碰撞筛选时间步长
# 论文正文需要单独列出的把手编号与时刻
PAPER_HANDLES = (0, 1, 51, 101, 151, 201, 223)
PAPER_TIMES = (-100.0, -50.0, 0.0, 50.0, 100.0)
# 对比图输出路径
COMPARISON_FIG = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "results",
    "figures",
    "problem4_turning_comparison.png",
)

def _quick_margin(g: dict, times: list[float] | None = None) -> float:
    """快速碰撞检查：返回指定时刻（默认调头窗口）内的最小裕量。

    Args:
        g: solve_s_curve 返回的几何 dict。
        times: 要检查的时刻列表；默认取调头窗口 t=0..30 步长 2 加最危险时刻。

    Returns:
        最小全局裕量（m），正为无碰撞。
    """
    if times is None:
        # 调头窗口聚焦检查：最危险时段密集 + 两侧稀疏
        times = list(range(14, 23)) + [0, 5, 10, 30]
    min_g = 1e9
    for t in times:
        pos = compute_dragon_at(float(t), g, need_speed=False)
        G, _, _ = global_margin(pos)
        min_g = min(min_g, G)
    return min_g


def _refine_local(ta0: float, tc0: float, orientation: int) -> tuple[float, float] | None:
    """以 (ta0, tc0) 为初值做纯几何网格收缩细化，返回更优切点或 None。

    只按弧长最小化（不查碰撞），快速得到局部极小；碰撞复核由调用方完成。

    Args:
        ta0, tc0: 初值切点参数。
        orientation: S 形朝向。

    Returns:
        (ta, tc)：细化后弧长更小的切点；若无可行细化返回 None。
    """
    ta, tc = ta0, tc0
    span = 0.05
    best_L = float("inf")
    best = None
    for _ in range(3):
        found = False
        ts = np.linspace(max(1e-6, ta - span), ta + span, 11)
        cs = np.linspace(max(1e-6, tc - span), tc + span, 11)
        for t1 in ts:
            for t2 in cs:
                r = turning_candidate(float(t1), float(t2), orientation)
                if r is not None and r[0] < best_L:
                    best_L = r[0]
                    best = (float(t1), float(t2))
                    found = True
        if not found:
            break
        ta, tc = best
        span *= 0.35
    if best is not None and best_L < turning_candidate(ta0, tc0, orientation)[0]:
        return best
    return None


def _search_orientation(
    ta0: float,
    tc0: float,
    orientation: int,
    delta: float,
    n_pert: int,
    L_opt: float,
    logger,
) -> dict | None:
    """在指定 S 形朝向分支的邻域内搜索更短的无碰撞解。

    步骤：
    1. 在 (ta0, tc0) 附近 delta 范围取 n_pert×n_pert 扰动网格，做几何筛选；
    2. 对弧长更短的候选做调头窗口碰撞复核，剔除碰撞方案；
    3. 对最短无碰撞候选做纯几何局部细化，细化后复核碰撞，碰撞则回退扰动网格解。

    Args:
        ta0, tc0: 参考切点参数（当前最优）。
        orientation: 要搜索的 S 形朝向。
        delta: 扰动半径（rad）。
        n_pert: 每维扰动网格点数。
        L_opt: 当前最优弧长（用于判定是否有改进）。
        logger: 日志记录器。

    Returns:
        该朝向分支内更优的无碰撞几何 dict；若未发现改进返回 None。
    """
    thetas = ta0 + delta * np.linspace(-1.0, 1.0, n_pert)
    phis = tc0 + delta * np.linspace(-1.0, 1.0, n_pert)
    L_min = L_opt
    best_ta, best_tc = ta0, tc0
    n_feasible = 0
    for ta in thetas:
        for tc in phis:
            r = turning_candidate(float(ta), float(tc), orientation)
            if r is None:
                continue
            n_feasible += 1
            L = r[0]
            if L < L_min - 1e-12:
                g_tmp = solve_s_curve(float(ta), float(tc), orientation)
                if g_tmp is None:
                    continue
                if _quick_margin(g_tmp) > 0.0:
                    L_min = L
                    best_ta, best_tc = float(ta), float(tc)
                    logger.info(
                        "    [%s] 发现更短无碰撞候选 (θa=%.6f, θc=%.6f) L=%.6f",
                        "PLUS" if orientation == ORIENTATION_PLUS else "MINUS",
                        best_ta, best_tc, L)
    if L_min >= L_opt - 1e-12:
        # 该朝向无改进，但仍返回 None
        return None
    # 局部细化
    refined = _refine_local(best_ta, best_tc, orientation)
    if refined is not None:
        best_ta, best_tc = refined
        g_tmp = solve_s_curve(best_ta, best_tc, orientation)
        if g_tmp is not None:
            if _quick_margin(g_tmp) > 0.0:
                L_min = g_tmp["L_S"]
                logger.info(
                    "    [%s] 局部细化后：L=%.9f（θa=%.9f, θc=%.9f）",
                    "PLUS" if orientation == ORIENTATION_PLUS else "MINUS",
                    L_min, best_ta, best_tc)
            else:
                logger.info(
                    "    [%s] 细化解碰撞，回退扰动网格候选 (θa=%.6f, θc=%.6f)",
                    "PLUS" if orientation == ORIENTATION_PLUS else "MINUS",
                    best_ta, best_tc)
    # 构造返回几何并复核
    g_new = solve_s_curve(best_ta, best_tc, orientation)
    if g_new is not None and _quick_margin(g_new) > 0.0:
        return g_new
    # 最终候选碰撞（边界附近），退回扰动网格中确认无碰撞的最短解
    logger.warning(
        "    [%s] 最终候选碰撞，退回扰动网格中最短无碰撞解",
        "PLUS" if orientation == ORIENTATION_PLUS else "MINUS")
    best_L2 = L_opt
    best2 = (ta0, tc0)
    for ta in thetas:
        for tc in phis:
            r = turning_candidate(float(ta), float(tc), orientation)
            if r is None or r[0] >= best_L2:
                continue
            g_tmp = solve_s_curve(float(ta), float(tc), orientation)
            if g_tmp is not None and _quick_margin(g_tmp) > 0.0:
                best_L2 = r[0]
                best2 = (float(ta), float(tc))
    g_fallback = solve_s_curve(best2[0], best2[1], orientation)
    return g_fallback if g_fallback is not None else None


def _search_other_orientation(
    ta0: float,
    tc0: float,
    orientation: int,
    L_opt: float,
    logger,
    wide_range: float = 0.5,
    n_wide: int = 15,
    n_refine: int = 25,
) -> dict | None:
    """在另一朝向分支做广域搜索，寻找更短的无碰撞解。

    由于另一朝向的最优解可能远离当前参考切点（超出小扰动半径），
    本函数以当前切点为中心在较大范围 [−wide_range, +wide_range] 内做
    粗网格，对弧长最短的前 n_refine 个候选做纯几何局部细化，再按细化后
    弧长升序复核碰撞，返回该朝向的最短无碰撞解。

    注意：细化后的解与原始几何弧长排序往往不一致——原始几何最短的候选
    细化后常落入碰撞区，而无碰撞的窄窗口最优需要覆盖足够多的细化候选
    才能采到（例如 MINUS 分支 L*≈1.287 的窄窗口来自原始弧长较大的
    (θa≈5.54, θc≈2.72) 附近候选）。因此 n_refine 需明显大于 10。

    Args:
        ta0, tc0: 参考切点参数（当前最优）。
        orientation: 要搜索的另一 S 形朝向。
        L_opt: 当前最优弧长。
        logger: 日志记录器。
        wide_range: 广域搜索半径（rad，默认 0.5）。
        n_wide: 每维广域网格点数（默认 15）。
        n_refine: 参与局部细化的候选数（默认 25）。

    Returns:
        该朝向更短的无碰撞几何 dict；若未发现返回 None。
    """
    name = "PLUS" if orientation == ORIENTATION_PLUS else "MINUS"
    # 广域粗网格
    tas = np.linspace(max(1e-6, ta0 - wide_range), ta0 + wide_range, n_wide)
    tcs = np.linspace(max(1e-6, tc0 - wide_range), tc0 + wide_range, n_wide)
    feasible: list[tuple[float, float, float]] = []  # (L, ta, tc)
    for ta in tas:
        for tc in tcs:
            r = turning_candidate(float(ta), float(tc), orientation)
            if r is not None:
                feasible.append((r[0], float(ta), float(tc)))
    if not feasible:
        logger.info("    [%s] 广域搜索无可行候选", name)
        return None
    feasible.sort(key=lambda x: x[0])
    # 对弧长最短的前 n_refine 个候选做纯几何细化，收集细化结果并按细化后
    # 弧长升序排列（细化可能显著改变弧长与排序）。
    refined_all: list[tuple[float, float, float]] = []  # (L, ta, tc)
    seen: set[tuple[float, float]] = set()
    for L, ta, tc in feasible[:n_refine]:
        refined = _refine_local(ta, tc, orientation)
        rta, rtc = (refined if refined is not None else (ta, tc))
        key = (round(rta, 7), round(rtc, 7))
        if key in seen:
            continue
        seen.add(key)
        g_tmp = solve_s_curve(rta, rtc, orientation)
        if g_tmp is not None:
            refined_all.append((g_tmp["L_S"], rta, rtc))
    if not refined_all:
        logger.info("    [%s] 广域搜索无细化结果", name)
        return None
    refined_all.sort(key=lambda x: x[0])
    # 按细化后弧长升序复核碰撞：第一个无碰撞者即该朝向最短无碰撞解。
    best_cand: tuple[float, float, float] | None = None
    for L, ta, tc in refined_all:
        if L >= L_opt - 1e-12:
            break  # 后续弧长更大，不可能更优
        g_tmp = solve_s_curve(ta, tc, orientation)
        if g_tmp is not None and _quick_margin(g_tmp) > 0.0:
            best_cand = (L, ta, tc)
            logger.info("    [%s] 广域搜索复核通过：细化候选 (θa=%.6f, θc=%.6f) L=%.6f",
                        name, ta, tc, L)
            break
    if best_cand is None:
        logger.info("    [%s] 广域搜索细化候选（L<当前最优）均碰撞，无更优无碰撞解", name)
        return None
    L, ta, tc = best_cand
    if L >= L_opt - 1e-12:
        return None
    logger.info("    [%s] 广域搜索发现更短无碰撞候选 (θa=%.6f, θc=%.6f) L=%.6f",
                name, ta, tc, L)
    g_new = solve_s_curve(ta, tc, orientation)
    if g_new is not None and g_new["L_S"] < L_opt - 1e-12 and _quick_margin(g_new) > 0.0:
        return g_new
    return None


def verify_local_optimality(
    g_opt: dict,
    logger,
    delta: float = 0.08,
    n_pert: int = 9,
) -> dict:
    """对最优切点 (θa, θc) 做小幅二维扰动，验证并改进全局最优（文档 8.4 节）。

    与单朝向版本的区别：**同时枚举两种 S 形朝向（ORIENTATION_PLUS 与
    ORIENTATION_MINUS）**：
    - 当前朝向：围绕最优切点做小扰动（delta 邻域）精细搜索；
    - 另一朝向：做广域搜索（wide_range 邻域），因为其最优解可能远离当前切点。

    最后比较两种朝向的结果，返回更短的无碰撞解；均无改进则返回原几何。
    该设计使最终最优不依赖粗搜索网格先命中哪个朝向分支。

    Args:
        g_opt: 最优配置几何 dict（solve_s_curve 返回）。
        logger: 日志记录器。
        delta: 当前朝向扰动半径（rad）。
        n_pert: 当前朝向每维扰动网格点数。

    Returns:
        更新后的最优几何 dict g_opt（若扰动发现更短无碰撞解则返回新解）。
    """
    logger.info("步骤 3.5：最优切点邻域扰动验证（文档 8.4 节，双朝向搜索）")
    ta0, tc0 = g_opt["theta_a"], g_opt["theta_c"]
    cur_ori = ORIENTATION_PLUS if g_opt["sgn1"] > 0 else ORIENTATION_MINUS
    other_ori = ORIENTATION_MINUS if cur_ori == ORIENTATION_PLUS else ORIENTATION_PLUS
    L_opt = g_opt["L_S"]

    best = (g_opt, L_opt)

    # 1) 当前朝向：小扰动精细搜索
    cand_cur = _search_orientation(ta0, tc0, cur_ori, delta, n_pert, L_opt, logger)
    if cand_cur is not None and cand_cur["L_S"] < best[1]:
        best = (cand_cur, cand_cur["L_S"])

    # 2) 另一朝向：广域搜索
    cand_other = _search_other_orientation(
        ta0, tc0, other_ori, best[1], logger)
    if cand_other is not None and cand_other["L_S"] < best[1]:
        best = (cand_other, cand_other["L_S"])

    if best[1] >= L_opt - 1e-12:
        logger.info("    验证通过：两种朝向邻域内弧长均不再下降，为局部最优 [OK]")
        return g_opt
    logger.info("    扰动发现更短无碰撞解（朝向 %s），返回改进后的最优配置",
                "PLUS" if best[0]["sgn1"] > 0 else "MINUS")
    return best[0]


# 碰撞约束最优配置（已验证全时段无碰撞）：切点参数 (θa, θc)
# 题目无圆弧角约束，此配置为无约束碰撞筛选 + 双朝向扰动验证的全局最短曲线
# L*≈1.287452 m（ORIENTATION_MINUS 分支）：
# （θa=5.602429, θc=2.656490，ORIENTATION_MINUS；弧1≈179.31°、弧2≈0.012°，
#   R=0.205686，minG=+0.000465 @ t=17，板凳对 1,9）。
# 该解由“最优切点邻域扰动验证（步骤 3.5）”的双朝向广域搜索自动发现：
# 400×400 网格碰撞筛选先命中 ORIENTATION_PLUS 分支 L*=1.291245（θa=5.753765,
#   θc=2.794193），随后扰动验证在 ORIENTATION_MINUS 分支广域搜索中经局部细化
#   收敛到更短的 L*=1.287452（原始几何候选 θa≈5.54, θc≈2.72 附近）。
# 注：网格密度只影响先命中的朝向分支，双朝向扰动验证保证了最终最优不依赖网格。
# 若 RUN_SCREENING=False，直接采用此配置，跳过耗时的碰撞筛选。
VERIFIED_OPTIMAL = (5.602429, 2.656490)
# 已验证最优配置的 S 形朝向（ORIENTATION_MINUS）
VERIFIED_OPTIMAL_ORIENTATION = ORIENTATION_MINUS
RUN_SCREENING = True  # 设为 True 时重新运行完整碰撞筛选


# 碰撞筛选参数
SCREEN_L_MAX = 1.6     # 粗筛扫描的调头曲线长度上限（m）
SCREEN_SUBSAMPLE = 3   # 粗筛候选子采样间隔
SCREEN_BISECT = 25     # 粗筛弦长二分迭代次数（低精度，加快筛选）
SCREEN_WORKERS = max(1, min(8, os.cpu_count() or 4))  # 并行筛选进程数
# 粗筛时刻（调头区，2s 步长以捕捉 t=8/10 附近的短暂碰撞）
SCREEN_TIMES = list(range(0, 42, 2))


def min_margin_over_range(
    g: dict,
    times: list[float] | None = None,
    step: float = TIME_STEP,
    bisect_iter: int = 45,
    logger=None,
) -> tuple[float, float, int, int]:
    """计算几何配置在指定时刻（或 t∈[T_START, T_END] 步长 step）内的最小碰撞裕量。

    Args:
        g: solve_s_curve 返回的几何 dict。
        times: 要检查的时刻列表；为 None 时用 [T_START, T_END] 步长 step。
        step: 时间步长（s），当 times 为 None 时生效。
        bisect_iter: 弦长方程二分迭代次数。
        logger: 日志记录器（可选）。

    Returns:
        (min_G, t_worst, i_worst, j_worst)：
        最小全局裕量及其对应时刻、最危险板凳对（1-based）。
    """
    if times is None:
        times = []
        t = T_START
        while t <= T_END + 1e-12:
            times.append(t)
            t += step
    min_g = 1e9
    worst = (times[0], 1, 1)
    for t in times:
        positions = compute_dragon_at(t, g, need_speed=False, bisect_iter=bisect_iter)
        G, i, j = global_margin(positions)
        if G < min_g:
            min_g = G
            worst = (t, i, j)
    if logger is not None:
        logger.info("    最小裕量 minG=%.6f m @ t=%.1f s（板凳对 %d, %d）",
                    min_g, worst[0], worst[1], worst[2])
    return min_g, worst[0], worst[1], worst[2]


def _screen_worker(payload: tuple) -> tuple:
    """多进程粗筛工作函数：计算单个候选在调头时刻的最小裕量。

    Args:
        payload: (idx, candidate, times, bisect_iter)
            idx: 候选在 candidates 列表中的下标；
            candidate: optimize_turning_curve 候选元组 (L, θa, θc, orientation, g)；
            times: 要检查的时刻列表；
            bisect_iter: 弦长二分迭代次数。

    Returns:
        (idx, L, θa, θc, min_G, t_worst, i_worst, j_worst)：
        候选下标、弧长、切点参数与最小裕量及其时刻、最危险板凳对。
    """
    idx, (L, ta, tc, _ori, g), times, bisect_iter = payload
    m, tw, iw, jw = min_margin_over_range(g, times=times, bisect_iter=bisect_iter)
    return idx, L, ta, tc, m, tw, iw, jw


def find_shortest_collision_free(candidates, logger) -> tuple:
    """从几何候选（按弧长升序）中找到最短无碰撞配置。

    由于碰撞裕量随弧长非单调（微小圆弧在中心附近必碰撞，较大圆弧可能安全），
    采用“子采样粗筛 + 全时段复核”的分阶段策略（文档 8.3 节）：

    1. 在 [0, SCREEN_L_MAX] 内按弧长子采样候选，用调头区 2s 时刻 + 低精度二分
       快速计算最小裕量，收集所有粗筛正裕量候选；
    2. 对粗筛通过的候选按弧长升序做全时段 t∈[-100,100]@1s 全精度复核，
       取首个全时段无碰撞者（非单调时跳过全检失败的候选继续检查）。

    Args:
        candidates: optimize_turning_curve 返回的候选列表
            [(L, θa, θc, orientation, g), ...]，已按 L 升序。
        logger: 日志记录器。

    Returns:
        (g_opt, L_star, min_G, t_worst, i_worst, j_worst, n_checked)：
        无碰撞候选的几何、弧长、最小裕量、最危险时刻与板凳对、检查数。

    Raises:
        RuntimeError: 未找到无碰撞候选。
    """
    logger.info("步骤 3：分阶段碰撞筛选（子采样粗筛 + 全时段复核）")
    Ls = np.array([c[0] for c in candidates])
    n_checked = 0
    i_scan = int(np.searchsorted(Ls, SCREEN_L_MAX, side="right"))

    # 阶段 3.1：子采样粗筛（多进程并行），收集粗筛正裕量候选
    rough_safe: list[int] = []
    screen_idxs = list(range(0, i_scan, SCREEN_SUBSAMPLE))
    tasks = [
        (idx, candidates[idx], SCREEN_TIMES, SCREEN_BISECT) for idx in screen_idxs
    ]
    with concurrent.futures.ProcessPoolExecutor(max_workers=SCREEN_WORKERS) as ex:
        for idx, L, ta, tc, m, tw, iw, jw in ex.map(_screen_worker, tasks):
            n_checked += 1
            if m > 0:
                rough_safe.append(idx)
                logger.info("    粗筛正裕量候选 idx=%d：L=%.4f（θa=%.4f, θc=%.4f）minG=%.4f",
                            idx, L, ta, tc, m)
    if not rough_safe:
        raise RuntimeError(f"在 L<{SCREEN_L_MAX} 内未找到粗筛正裕量候选")

    # 阶段 3.2：对粗筛通过候选按弧长升序全时段复核
    for idx in rough_safe:
        L, ta, tc, ori, g = candidates[idx]
        n_checked += 1
        m_full, tw, iw, jw = min_margin_over_range(g, step=TIME_STEP)
        logger.info("    候选 L=%.4f (θa=%.4f, θc=%.4f) 全时段 minG=%.6f @ t=%.1f（%d,%d）",
                    L, ta, tc, m_full, tw, iw, jw)
        if m_full > 0:
            logger.info("    找到首个全时段无碰撞候选：L*=%.6f (θa=%.6f, θc=%.6f, R=%.6f)",
                        L, ta, tc, g["R"])
            return g, L, m_full, tw, iw, jw, n_checked
    raise RuntimeError(f"检查了 {n_checked} 个候选仍未找到全时段无碰撞方案")


def compute_all_times(g: dict, logger) -> tuple[np.ndarray, np.ndarray]:
    """计算 t=T_START..T_END 全部时刻的把手位置与速度。

    Args:
        g: solve_s_curve 返回的几何 dict。
        logger: 日志记录器。

    Returns:
        (positions_all, speeds_all)：
            positions_all: shape (n_times, TOTAL_HANDLES, 2)；
            speeds_all: shape (n_times, TOTAL_HANDLES)。
    """
    n_times = int(round((T_END - T_START) / TIME_STEP)) + 1
    times = T_START + TIME_STEP * np.arange(n_times)
    positions_all = np.zeros((n_times, TOTAL_HANDLES, 2))
    speeds_all = np.zeros((n_times, TOTAL_HANDLES))
    for k, t in enumerate(times):
        positions, speeds = compute_dragon_at(float(t), g)
        positions_all[k] = positions
        speeds_all[k] = speeds
        if k % 50 == 0:
            logger.info("    t=%5.0f s 计算完成", t)
    return positions_all, speeds_all


def write_result4(
    positions_all: np.ndarray,
    speeds_all: np.ndarray,
    output_path: str,
) -> None:
    """将全部把手位置与速度写入 result4.xlsx（覆盖旧文件）。

    Args:
        positions_all: (n_times, 224, 2)。
        speeds_all: (n_times, 224)。
        output_path: 输出文件路径。
    """
    wb = openpyxl.load_workbook(P4_RESULT_TEMPLATE_PATH)
    ws_pos = wb["位置"]
    ws_spd = wb["速度"]
    n_times = positions_all.shape[0]
    for k in range(n_times):
        col = k + 2  # 第 1 列为标签，时间列从第 2 列开始
        for i in range(TOTAL_HANDLES):
            x, y = positions_all[k, i]
            ws_pos.cell(row=2 + 2 * i, column=col).value = round(float(x), 6)
            ws_pos.cell(row=3 + 2 * i, column=col).value = round(float(y), 6)
            ws_spd.cell(row=2 + i, column=col).value = round(float(speeds_all[k, i]), 6)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


def log_paper_table(positions_all, speeds_all, times, logger) -> None:
    """记录论文正文所需的关键时刻与把手位置、速度。"""
    idx = {t: int(round((t - T_START) / TIME_STEP)) for t in PAPER_TIMES}
    for t in PAPER_TIMES:
        k = idx[t]
        logger.info("---- t=%5.0f s ----", t)
        for i in PAPER_HANDLES:
            x, y = positions_all[k, i]
            v = speeds_all[k, i]
            name = "龙头" if i == 0 else (f"龙尾后" if i == 223 else f"龙身第{i}节")
            logger.info("    %s: (%.6f, %.6f), v=%.6f m/s", name, x, y, v)


def verify_model(g: dict, logger) -> None:
    """运行问题四模型检验（文档第 14 节）。

    Args:
        g: solve_s_curve 返回的几何 dict。
        logger: 日志记录器。
    """
    logger.info("步骤 6：模型检验")
    # 6.1 半径比检验：两段圆弧半径应为 R1=2R、R2=R，满足 R1=2R2
    # （该关系由构造保证，此处作为一致性检查记录）
    logger.info("  半径比 R1/R2 = %.4f（应=2.0）", 2.0 * g["R"] / g["R"])
    # 6.2 区域约束：两段圆弧须位于调头空间圆内（|P| ≤ ρ=%.2f m）
    arc1, arc2 = s_curve_points(g, n1=300, n2=300)
    pts = np.vstack([arc1, arc2])
    max_r = float(np.max(np.linalg.norm(pts, axis=1)))
    logger.info("  调头空间约束 max|P| = %.4f m（应≤ρ=%.2f m）", max_r, TURN_RADIUS)
    # 6.3 相切与连续性：|Γ'(s)|=1（弧长参数化）
    max_dev = 0.0
    for s in [-80.0, -5.0, g["L1"] * 0.5, g["L1"] + 0.5, g["L_S"] + 10.0, 80.0]:
        h = 1e-6
        v = (path_point(s + h, g) - path_point(s - h, g)) / (2 * h)
        max_dev = max(max_dev, abs(float(np.linalg.norm(v)) - 1.0))
    logger.info("  弧长参数化最大 |Γ'|-1 = %.3e（应≈0）", max_dev)
    # 6.4 刚性距离与速度约束（抽查若干时刻）
    dists = handle_distances()
    max_chord = 0.0
    max_rig = 0.0
    for t in (T_START, 0.0, 30.0, T_END):
        positions, speeds, s_params = compute_dragon_at(
            t, g, need_speed=True, return_s_params=True)
        # 带符号的弧长变化率 ṡ（式 40），速度向量 v_i = ṡ_i·τ_i
        sdot = np.zeros(TOTAL_HANDLES)
        sdot[0] = 1.0
        for i in range(1, TOTAL_HANDLES):
            max_chord = max(max_chord, abs(
                float(np.linalg.norm(positions[i] - positions[i - 1])) - dists[i]))
            tau_prev = path_tangent(s_params[i - 1], g)
            tau_i = path_tangent(s_params[i], g)
            dP = positions[i] - positions[i - 1]
            sdot[i] = sdot[i - 1] * (np.dot(dP, tau_prev) / np.dot(dP, tau_i))
            max_rig = max(
                max_rig,
                abs(float(np.dot(dP, sdot[i] * tau_i - sdot[i - 1] * tau_prev))),
            )
    logger.info("  最大刚性距离误差 = %.3e（应≈0）", max_chord)
    logger.info("  最大速度刚性约束残差 = %.3e（应≈0）", max_rig)


def main() -> None:
    t_start_all = time.time()
    logger = setup_logger(__name__, script_name="solve_problem4")
    logger.info("问题四：S 形调头曲线优化与运动计算开始")

    # 1. 基准配置
    logger.info("步骤 1：求解基准配置（两圆弧与调头空间圆内切）")
    state, f, g0 = solve_given_configuration(16.5715, 16.6169, ORIENTATION_MINUS)
    L0 = g0["L_S"]
    logger.info("    基准：θa=%.9f，θc=%.9f，R=%.6f，L0=%.6f m",
                state[0], state[1], g0["R"], L0)

    # 2. 几何优化（仅重新筛选时需要；否则使用已验证配置）
    candidates = None
    if RUN_SCREENING:
        logger.info("步骤 2：二维几何粗搜索 + 多起点局部细化")
        best_geom, candidates = optimize_turning_curve(
            n_grid=400, top_k=5, zoom_rounds=5, zoom_grid=21,
        )
        L_geom, ta_g, tc_g, ori_g, g_geom = best_geom
        logger.info("    几何最优：L=%.6f（θa=%.6f，θc=%.6f，R=%.6f）",
                    L_geom, ta_g, tc_g, g_geom["R"])
        logger.info("    可行几何候选数：%d", len(candidates))
    else:
        logger.info("步骤 2：跳过几何优化（采用已验证最优配置，如需重新筛选请设 RUN_SCREENING=True）")

    # 3. 碰撞约束最优配置
    if RUN_SCREENING:
        logger.info("步骤 3：重新运行碰撞筛选（文档 8.3 节）")
        g_opt, L_star, m_opt, tw, iw, jw, n_checked = find_shortest_collision_free(
            candidates, logger
        )
    else:
        logger.info("步骤 3：采用已验证的无碰撞最优配置")
        ta_v, tc_v = VERIFIED_OPTIMAL
        g_opt = solve_s_curve(ta_v, tc_v, VERIFIED_OPTIMAL_ORIENTATION)
        if g_opt is None:
            raise RuntimeError("已验证配置无法求解")
        L_star = g_opt["L_S"]
        m_opt, tw, iw, jw = min_margin_over_range(g_opt, step=TIME_STEP)
        logger.info("    配置：θa=%.9f，θc=%.9f，R=%.6f，L*=%.6f m",
                    ta_v, tc_v, g_opt["R"], L_star)
        logger.info("    全时段复核 minG=%.6f m @ t=%.1f（板凳对 %d, %d）",
                    m_opt, tw, iw, jw)
    dL = L0 - L_star
    eta = 100.0 * dL / L0
    logger.info("优化结果：L*=%.6f m，ΔL=%.6f m，缩短比例 η=%.2f%%", L_star, dL, eta)
    logger.info("最优配置：θa=%.9f，θc=%.9f，R=%.6f，2R=%.6f，最小裕量=%.6f m",
                g_opt["theta_a"], g_opt["theta_c"], g_opt["R"],
                2 * g_opt["R"], m_opt)

    # 3.5 全局性验证：最优切点邻域扰动（文档 8.4 节）
    g_opt = verify_local_optimality(g_opt, logger)
    L_star = g_opt["L_S"]
    m_opt, tw, iw, jw = min_margin_over_range(g_opt, step=TIME_STEP)
    logger.info("    最终采用：L*=%.9f m（θa=%.9f, θc=%.9f），minG=%.6f @ t=%.1f（%d,%d）",
                L_star, g_opt["theta_a"], g_opt["theta_c"], m_opt, tw, iw, jw)
    dL = L0 - L_star
    eta = 100.0 * dL / L0
    logger.info("优化结果：L*=%.6f m，ΔL=%.6f m，缩短比例 η=%.2f%%", L_star, dL, eta)
    logger.info("最优配置：θa=%.9f，θc=%.9f，R=%.6f，2R=%.6f，最小裕量=%.6f m",
                g_opt["theta_a"], g_opt["theta_c"], g_opt["R"],
                2 * g_opt["R"], m_opt)

    # 4. 计算全程位置与速度（采用最优配置）
    logger.info("步骤 4：计算 t=%g..%g s 全部把手位置与速度（最优配置）", T_START, T_END)
    positions_all, speeds_all = compute_all_times(g_opt, logger)

    # 5. 写入 result4.xlsx
    logger.info("步骤 5：写入 result4.xlsx")
    write_result4(positions_all, speeds_all, P4_RESULT_PATH)
    logger.info("    已保存：%s", P4_RESULT_PATH)

    # 6. 模型检验
    verify_model(g_opt, logger)

    # 7. 论文表格
    log_paper_table(positions_all, speeds_all,
                    T_START + TIME_STEP * np.arange(positions_all.shape[0]), logger)

    # 8. 绘图
    logger.info("步骤 8：绘制基准/最优调头曲线对比图")
    plot_turning_comparison(g0, g_opt, COMPARISON_FIG)
    logger.info("    已保存：%s", COMPARISON_FIG)

    logger.info("完成，总耗时 %.1f s", time.time() - t_start_all)


if __name__ == "__main__":
    main()
