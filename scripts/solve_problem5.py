"""问题五求解脚本：龙头最大行进速度优化。

运行方式:
    python scripts/solve_problem5.py

功能:
    1. 加载问题四最优配置（S 形调头曲线），建立弧长参数化路径；
    2. 读取 ``result4.xlsx`` 单位速度（1 m/s）下的 224×201 速度结果，
       直接取表格最大值作为第一候选（文档第 10.2 节）；
    3. 宽范围粗扫 + 局部加密 + 峰值细化，求全程最大速度放大系数
       ``Λ_max``、临界构型 ``s0*``、临界把手 ``i*``（文档第 10.3~10.4 节）；
    4. 由 ``v_h^max = 2/Λ_max`` 直接计算龙头最大恒定速度，不枚举龙头速度；
    5. 运行模型检验：刚性距离、速度约束、线性缩放、最优边界、
       网格收敛、搜索区间扩展、奇异分母监测（文档第 15 节）；
    6. 输出结果表格 ``results/tables/problem5_lambda_curve.xlsx``、
       数值汇总 ``results/tables/problem5_summary.json`` 与三张图；
    7. 日志写入 ``results/logs/solve_problem5_YYYY-MM-DD_HHMMSS.log``（不覆盖旧文件）。
"""

import json
import os
import time

import numpy as np
import openpyxl

from src.config import (
    P4_RESULT_PATH,
    P5_CONFIG_FIG,
    P5_CONFIG_FIG,
    P5_FINE_STEP,
    P5_LAMBDA_FIG,
    P5_LAMBDA_TABLE,
    P5_REFINE_RADIUS,
    P5_SPEEDS_FIG,
    P5_SUMMARY_PATH,
    P5_SPEED_LIMIT,
    TOTAL_HANDLES,
)
from src.models.problem4 import compute_dragon_at
from src.models.problem5 import (
    build_summary,
    fine_lambda_curve,
    load_optimal_geometry,
    solve_max_amplification,
    verify_all,
)
from src.utils.logger import setup_logger
from src.visualization.problem5_plots import (
    plot_critical_configuration,
    plot_critical_speeds,
    plot_lambda_curve,
)


def _parse_time_label(label) -> float:
    """解析 Excel 时间表头（如 '-100 s'、'-99' 或数值）为秒数。

    Args:
        label: 表头单元格内容。

    Returns:
        对应的时刻（s）；无法解析时返回 0.0。
    """
    if isinstance(label, (int, float)):
        return float(label)
    if isinstance(label, str):
        s = label.strip()
        for suffix in ("s", "秒"):
            if s.endswith(suffix):
                s = s[: -len(suffix)]
        try:
            return float(s)
        except ValueError:
            return 0.0
    return 0.0


def read_result4_max_speed(path: str) -> tuple[float, float, int]:
    """读取 ``result4.xlsx`` 速度表的最大值及其（时刻, 把手）位置。

    ``result4.xlsx`` 的“速度”工作表：第 1 行为时间表头，第 1 列为把手标签，
    数据区为 224 行 × 201 列的速度值（龙头速度 1 m/s，故数值即放大系数）。

    Args:
        path: result4.xlsx 的绝对路径。

    Returns:
        ``(max_speed, t_at_max, handle_idx)``：
        表格最大值、对应时刻（s）、对应把手编号。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[-1]]
    rows = list(ws.iter_rows(values_only=True))
    best = (0.0, 0.0, 0)
    for r in range(1, len(rows)):
        for c in range(1, len(rows[r])):
            v = rows[r][c]
            if isinstance(v, (int, float)) and float(v) > best[0]:
                best = (float(v), _parse_time_label(rows[0][c]), r - 1)
    return best


def write_lambda_table(
    s0_grid: np.ndarray,
    lambda_curve: np.ndarray,
    lambda_star: np.ndarray,
    s0_star: float,
    i_star: int,
    v_h_max: float,
    output_path: str,
) -> None:
    """将 ``Λ(s)`` 扫描曲线与临界构型速度分布写入 Excel（覆盖旧文件）。

    Args:
        s0_grid: 粗扫龙头弧长网格。
        lambda_curve: 各网格点上的 ``Λ(s0)``。
        lambda_star: 临界构型下全部把手放大系数。
        s0_star: 临界龙头弧长位置（m）。
        i_star: 临界把手编号。
        v_h_max: 龙头最大恒定速度（m/s）。
        output_path: 输出文件路径。
    """
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Lambda曲线"
    ws1.append(["龙头弧长 s0 (m)", "最大放大系数 Lambda(s0)"])
    for s, lam in zip(s0_grid, lambda_curve):
        ws1.append([round(float(s), 6), round(float(lam), 9)])

    ws2 = wb.create_sheet("临界构型")
    ws2.append(["把手编号 i", "放大系数 lambda_i", "速度 v_i (m/s)"])
    for i in range(lambda_star.size):
        ws2.append([
            i,
            round(float(lambda_star[i]), 9),
            round(float(lambda_star[i]) * v_h_max, 9),
        ])
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


def log_paper_result(result: dict, checks: dict, logger) -> None:
    """记录论文正文（文档第 14 节）所需的关键数值。"""
    g = result["g"]
    s0_star = result["s0_star"]
    i_star = result["i_star"]
    v_h_max = result["v_h_max"]
    logger.info("---- 问题五论文结果 ----")
    logger.info("全程最大放大系数 Λ_max      = %.9f", result["Lambda_max"])
    logger.info("龙头最大恒定速度 v_h^max     = %.9f m/s", v_h_max)
    logger.info("临界把手编号 i*              = %d（第 %d 节龙身）", i_star, i_star)
    logger.info("临界龙头弧长位置 s*          = %.6f m", s0_star)
    logger.info("临界时刻 t* = s*/v_h^max     = %.6f s", s0_star / v_h_max)
    logger.info("临界构型下临界把手速度       = %.9f m/s（应=2）",
                result["lambda_star"][i_star] * v_h_max)
    logger.info("最小几何分母 D_min            = %.4e m", result["d_min"])
    logger.info("几何：θa=%.9f, θc=%.9f, R=%.9f, L1=%.9f, L2=%.9f, L_S=%.9f",
                g["theta_a"], g["theta_c"], g["R"], g["L1"], g["L2"], g["L_S"])
    logger.info("表格(1s)最大 Λ=%.9f @ t=%g s，局部加密后 Λ_max=%.9f",
                result["grid_lambda_max"],
                result["s0_grid"][int(np.argmax(result["lambda_curve"]))],
                result["Lambda_max"])
    logger.info("速度上限 = %.1f m/s", P5_SPEED_LIMIT)
    # 检验摘要
    logger.info("---- 问题五检验摘要 ----")
    logger.info("刚性距离最大误差      = %.3e", checks["max_chord_error"][0])
    logger.info("速度约束最大残差      = %.3e", checks["max_speed_residual"][0])
    logger.info("线性缩放最大偏差      = %.3e", checks["linear_scaling_error"])
    logger.info("网格 vs 加密相对差    = %.4f%%", checks["grid_vs_fine_rel_diff"])
    logger.info("搜索区间扩展最大 Λ    = %.6f @ s0=%.2f", *checks["extend_scan"])


def main() -> None:
    t_start = time.time()
    logger = setup_logger(__name__, script_name="solve_problem5")
    logger.info("问题五：龙头最大行进速度优化开始")

    # 1. 加载问题四最优配置
    logger.info("步骤 1：加载问题四最优配置")
    g = load_optimal_geometry()
    logger.info("    θa=%.6f，θc=%.6f，R=%.6f，L1=%.6f，L2=%.6f，L_S=%.6f m",
                g["theta_a"], g["theta_c"], g["R"], g["L1"], g["L2"], g["L_S"])

    # 2. 读取 result4.xlsx 单位速度结果（文档第 10.2 节）
    logger.info("步骤 2：读取 result4.xlsx 单位速度结果（1 s 间隔）")
    grid_lam, grid_t, grid_i = read_result4_max_speed(P4_RESULT_PATH)
    logger.info("    表格最大值 Λ=%.6f @ t=%g s，把手 %d（第 %d 节）",
                grid_lam, grid_t, grid_i, grid_i)

    # 3. 主扫描（粗扫 + 局部加密 + 峰值细化）
    result = solve_max_amplification(g, logger)

    # 4. 模型检验（文档第 15 节）
    checks = verify_all(result, logger)

    # 5. 写结果表格与汇总
    logger.info("步骤 6：写结果表格与汇总")
    write_lambda_table(
        result["s0_grid"], result["lambda_curve"],
        result["lambda_star"], result["s0_star"],
        result["i_star"], result["v_h_max"], P5_LAMBDA_TABLE,
    )
    logger.info("    已保存：%s", P5_LAMBDA_TABLE)
    summary = build_summary(result, checks)
    os.makedirs(os.path.dirname(P5_SUMMARY_PATH), exist_ok=True)
    with open(P5_SUMMARY_PATH, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    logger.info("    已保存：%s", P5_SUMMARY_PATH)

    # 6. 绘图
    logger.info("步骤 7：绘图")
    # 最大值附近精细采样（供局部放大子图，粗扫 1 m 步长不足以展示 0.2 m 宽的峰值）
    fine_s0, fine_lambda = fine_lambda_curve(
        g, result["s0_star"], P5_REFINE_RADIUS, P5_FINE_STEP
    )
    plot_lambda_curve(
        result["s0_grid"], result["lambda_curve"],
        result["s0_star"], result["Lambda_max"], P5_LAMBDA_FIG,
        fine_s0=fine_s0, fine_lambda=fine_lambda,
    )
    logger.info("    已保存：%s（含最大值附近局部放大）", P5_LAMBDA_FIG)
    plot_critical_speeds(
        result["lambda_star"], result["i_star"], result["v_h_max"],
        result["s0_star"], P5_SPEED_LIMIT, P5_SPEEDS_FIG,
    )
    logger.info("    已保存：%s", P5_SPEEDS_FIG)
    positions, _ = compute_dragon_at(result["s0_star"], g)
    plot_critical_configuration(
        g, positions, result["i_star"], result["s0_star"], P5_CONFIG_FIG,
    )
    logger.info("    已保存：%s", P5_CONFIG_FIG)

    # 7. 论文表格
    log_paper_result(result, checks, logger)

    logger.info("完成，总耗时 %.1f s", time.time() - t_start)


if __name__ == "__main__":
    main()
