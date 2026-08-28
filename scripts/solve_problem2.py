"""问题二求解脚本：板凳龙盘入终止时刻的碰撞检测。

运行方式:
    python scripts/solve_problem2.py

功能:
    1. 一秒步长粗搜索 + 局部加密扫描 + 二分法求首次接触时刻 t*；
    2. 在 t* 处计算全部 224 个把手的位置与速度；
    3. 运行模型检验（几何尺寸、终止时刻左右状态、接触对交叉验证）；
    4. 读取模板 data/result2.xlsx，将结果保存到 results/tables/result2.xlsx
       （不修改模板，覆盖旧输出）；
    5. 日志写入 results/logs/solve_problem2_YYYY-MM-DD_HHMMSS.log（不覆盖旧文件）。

性能说明:
    - 碰撞检测使用 numpy 对 24531 对非相邻板凳批量向量化（src/models/collision.py）；
    - 搜索阶段仅计算位置、不计算速度（compute_dragon_at(need_speed=False)）；
    - 粗搜索定位首次接触后，对最后 20 s 做 0.05 s 加密扫描，消除一秒采样
      可能漏掉短暂碰撞的风险（对应文档第 10.3 节）。
"""

import os
import time

import openpyxl

from src.config import (
    G_TABLE_PATH,
    RESULT2_PATH,
    RESULT2_TEMPLATE_PATH,
    TOTAL_HANDLES,
)
from src.models.collision import (
    global_margin,
    verify_geometry,
    verify_pair_independent,
)
from src.models.spiral_dragon import compute_dragon_at
from src.utils.logger import setup_logger
from src.visualization.collision_plots import plot_global_margin

# 二分收敛阈值（s），见文档式(32)
EPS_T = 1e-8
# 终止时刻左右状态验证步长（s），见文档式(36)
DELTA = 1e-6
# 终止时刻裕量容许误差（m），见文档式(37)
EPS_G = 1e-6
# 加密扫描步长与窗口（s），用于消除一秒粗搜索的漏检风险
FINE_STEP = 0.05
FINE_WINDOW = 20.0
# 粗搜索时间安全上限（s）
COARSE_LIMIT = 1000.0
# 论文正文需要单独列出的把手编号
PAPER_HANDLES = (0, 1, 51, 101, 151, 201, 223)


def margin_at(t: float, logger=None) -> tuple[float, int, int]:
    """计算时刻 t 的全局碰撞裕量及最危险板凳对（仅算位置，不算速度）。

    Args:
        t: 时间（s）。
        logger: 日志记录器，用于输出每个采样点的轨迹（可选）。

    Returns:
        (G, i_star, j_star)：全局裕量（m）与最危险板凳对（1-based 编号）。
    """
    positions, *_ = compute_dragon_at(t, need_speed=False)
    G, i_star, j_star = global_margin(positions)
    if logger is not None:
        logger.info("    t=%8.3f s  G=%.6f  最危险板凳对=(%d, %d)", t, G, i_star, j_star)
    return G, i_star, j_star


def coarse_search(logger) -> tuple[float, list[tuple[float, float, int, int]]]:
    """一秒步长粗搜索，返回首次满足 G<=0 的整数时刻及轨迹。

    Args:
        logger: 日志记录器。

    Returns:
        (t_rough, trace)：t_rough 为首次整数时刻 G<=0；
        trace 为 (t, G, i, j) 轨迹列表。
    """
    logger.info("步骤 1：一秒步长粗搜索")
    trace: list[tuple[float, float, int, int]] = []
    t = 0.0
    G, i, j = margin_at(t)
    trace.append((t, G, i, j))
    if G <= 0:
        logger.warning("初始时刻 G=%.6f<=0，直接判定 t*=0", G)
        return 0.0, trace
    while True:
        t += 1.0
        G, i, j = margin_at(t)
        trace.append((t, G, i, j))
        if G <= 0:
            logger.info("粗搜索在 t=%d s 首次检测到 G<=0（G=%.6f）", int(t), G)
            return t, trace
        if t >= COARSE_LIMIT:
            raise RuntimeError(f"超过 {COARSE_LIMIT:.0f} s 仍未检测到碰撞")


def fine_scan(t0: float, t1: float, logger) -> tuple[float, float, int]:
    """在 [t0, t1] 内以 FINE_STEP 步长加密扫描。

    用于消除一秒粗搜索可能漏掉采样点之间短暂碰撞的风险（文档 10.3 节）。

    Args:
        t0: 扫描起点（s）。
        t1: 扫描终点（s）。
        logger: 日志记录器。

    Returns:
        (last_safe, first_contact, n_points, points)：
        - last_safe: 最后一个 G>0 的加密时刻；
        - first_contact: 第一个 G<=0 的加密时刻（若存在，否则为 t1）；
        - n_points: 扫描点数；
        - points: 加密扫描点列表 [(t, G), ...]，用于局部放大绘图。
    """
    logger.info("步骤 2：局部加密扫描 [%.2f, %.2f] s，步长 %.3f s", t0, t1, FINE_STEP)
    last_safe = t0
    first_contact: float | None = None
    n_points = 0
    points: list[tuple[float, float]] = []
    t = t0
    while t <= t1 + 1e-12:
        G, i, j = margin_at(t)
        n_points += 1
        points.append((t, G))
        if G > 0:
            last_safe = t
        elif first_contact is None:
            first_contact = t
        t += FINE_STEP
    if first_contact is None:
        # t1 处应已 G<=0，此处兜底
        first_contact = t1
    logger.info(
        "加密扫描完成：%d 个采样点，最后安全时刻 %.4f s，首次接触 %.4f s",
        n_points, last_safe, first_contact,
    )
    return last_safe, first_contact, n_points, points


def bisect_search(tL: float, tR: float, logger) -> float:
    """在 [tL, tR] 内二分求首次接触时刻（文档第 10.2 节）。

    前提：G(tL)>0（安全侧），G(tR)<=0（接触/重叠侧）。

    Args:
        tL: 安全侧时刻（s）。
        tR: 接触侧时刻（s）。
        logger: 日志记录器。

    Returns:
        二分得到的终止时刻 t*（取右端点）。
    """
    logger.info("步骤 3：二分法精细搜索，收敛阈值 ε_t=%.0e s", EPS_T)
    iterations = 0
    while tR - tL >= EPS_T:
        tM = 0.5 * (tL + tR)
        GM, i, j = margin_at(tM)
        iterations += 1
        if GM > 0:
            tL = tM
        else:
            tR = tM
    logger.info("二分完成：%d 次迭代，t* = %.9f s", iterations, tR)
    return tR


def verify_termination(t_star: float, logger) -> None:
    """运行终止时刻左右状态检验（文档 15.2 节）。

    Args:
        t_star: 终止时刻（s）。
        logger: 日志记录器。
    """
    logger.info("步骤 4：终止时刻左右状态检验（δ=%.0e s）", DELTA)
    G0, i0, j0 = margin_at(t_star)
    Gm, im, jm = margin_at(t_star - DELTA)
    Gp, ip, jp = margin_at(t_star + DELTA)
    ok_left = Gm > 0
    ok_right = Gp < 0
    ok_zero = abs(G0) < EPS_G
    logger.info("  G(t*-δ)=%.3e (应>0, %s)", Gm, "通过" if ok_left else "失败")
    logger.info("  G(t*)=%.3e (应|·|<%.0e, %s)", G0, EPS_G, "通过" if ok_zero else "失败")
    logger.info("  G(t*+δ)=%.3e (应<0, %s)", Gp, "通过" if ok_right else "失败")
    if not (ok_left and ok_right and ok_zero):
        raise RuntimeError("终止时刻左右状态检验未通过")


def save_g_table(
    trace: list[tuple[float, float, int, int]],
    fine_points: list[tuple[float, float]],
    output_path: str,
) -> None:
    """将全局碰撞裕量 G(t) 数据保存到 xlsx 表格（覆盖旧文件）。

    两个工作表：
        - “粗搜索”：一秒步长采样，列含时间、全局裕量、最危险板凳对；
        - “加密扫描”：终止时刻附近 0.05 s 加密点（若提供）。

    Args:
        trace: 粗搜索轨迹，元素为 (t, G, i_star, j_star)。
        fine_points: 加密扫描点列表 [(t, G), ...]。
        output_path: 输出 xlsx 路径。
    """
    wb = openpyxl.Workbook()
    ws_coarse = wb.active
    ws_coarse.title = "粗搜索"
    ws_coarse.append(
        ["时间 t (s)", "全局裕量 G(t) (m)", "最危险板凳对 i", "最危险板凳对 j"]
    )
    for t, G, i, j in trace:
        ws_coarse.append([round(t, 6), round(G, 6), i, j])
    if fine_points:
        ws_fine = wb.create_sheet("加密扫描")
        ws_fine.append(["时间 t (s)", "全局裕量 G(t) (m)"])
        for t, G in fine_points:
            ws_fine.append([round(t, 6), round(G, 6)])
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


def fill_result2(
    template_path: str, output_path: str,
    positions: list, speeds: list,
) -> None:
    """读取模板，将终止时刻各把手位置与速度填入并保存。

    模板结构（data/result2.xlsx，Sheet1）：
        - 第 1 行为表头：横坐标x (m)、纵坐标y (m)、速度 (m/s)；
        - 把手 i 对应第 i+2 行（i=0 龙头，i=1..221 龙身，i=222 龙尾，i=223 龙尾(后)）；
        - 数值保留 6 位小数。

    Args:
        template_path: 模板文件路径（只读不修改）。
        output_path: 结果输出路径（results/tables/result2.xlsx）。
        positions: 224 个把手中心坐标列表。
        speeds: 224 个把手速度大小列表。
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Sheet1"]
    for i in range(TOTAL_HANDLES):
        row = i + 2
        ws.cell(row=row, column=2).value = round(positions[i][0], 6)
        ws.cell(row=row, column=3).value = round(positions[i][1], 6)
        ws.cell(row=row, column=4).value = round(speeds[i], 6)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


def main() -> None:
    logger = setup_logger(__name__, script_name="solve_problem2")
    logger.info("=" * 60)
    logger.info("问题二：板凳龙盘入终止时刻的碰撞检测")
    logger.info("方法：分离轴定理全局裕量 + 一秒粗搜索 + 加密扫描 + 二分")
    logger.info("把手总数：%d，非相邻板凳对数：C(223,2)-222 = 24531", TOTAL_HANDLES)
    logger.info("=" * 60)

    t_total = time.perf_counter()

    # ---- 步骤 1：一秒粗搜索 ----
    t_rough, trace = coarse_search(logger)

    # ---- 步骤 2：局部加密扫描，消除漏检风险 ----
    win_start = max(0.0, t_rough - FINE_WINDOW)
    last_safe, first_contact, _, fine_points = fine_scan(win_start, t_rough, logger)
    tL, tR = last_safe, first_contact
    if tR - tL <= 0:
        tL = max(0.0, tR - FINE_STEP)

    # ---- 步骤 3：二分法 ----
    t_star = bisect_search(tL, tR, logger)

    # ---- 步骤 4：终止时刻左右状态检验 ----
    verify_termination(t_star, logger)

    # ---- 步骤 5：终止时刻高精度重算全部把手位置与速度 ----
    logger.info("步骤 5：在 t* 处高精度计算全部把手位置与速度")
    positions, speeds, thetas, theta_dot, velocities = compute_dragon_at(
        t_star, need_speed=True
    )

    # ---- 步骤 6：模型检验 ----
    logger.info("步骤 6：模型检验")
    geo = verify_geometry(positions)
    for key, val in geo.items():
        logger.info("  %s: %.3e", key, val)
    G_star, i_star, j_star = global_margin(positions)
    colliding, detail = verify_pair_independent(positions, i_star, j_star)
    logger.info("  终止时刻全局裕量 G(t*)=%.3e", G_star)
    logger.info("  最危险板凳对=(%d, %d)，交叉验证：%s（%s）",
                i_star, j_star, "发生接触" if colliding else "未检测到接触", detail)

    # ---- 步骤 7：写入 result2.xlsx ----
    logger.info("步骤 7：写入结果文件")
    logger.info("  读取模板：%s", RESULT2_TEMPLATE_PATH)
    logger.info("  输出结果：%s", RESULT2_PATH)
    fill_result2(RESULT2_TEMPLATE_PATH, RESULT2_PATH, positions, speeds)

    # ---- 输出论文正文需要的把手结果 ----
    logger.info("=" * 60)
    logger.info("论文正文所需把手结果（t*=%.6f s）：", t_star)
    logger.info("  编号        横坐标x (m)    纵坐标y (m)     速度 (m/s)")
    for idx in PAPER_HANDLES:
        logger.info("  %4d   %14.6f   %14.6f   %12.6f",
                    idx, positions[idx][0], positions[idx][1], speeds[idx])

    # ---- 粗搜索轨迹抽样（展示 G 非单调性） ----
    logger.info("=" * 60)
    logger.info("粗搜索 G(t) 轨迹抽样（每 %d s 一点，最危险板凳对）：",
                max(1, len(trace) // 40))
    step = max(1, len(trace) // 40)
    for k in range(0, len(trace), step):
        tt, GG, ii, jj = trace[k]
        logger.info("  t=%6.1f s  G=%+.6f  pair=(%d, %d)", tt, GG, ii, jj)

    # ---- 保存 G(t) 数据表格 ----
    logger.info("步骤 8：保存 G(t) 数据表格")
    save_g_table(trace, fine_points, G_TABLE_PATH)
    logger.info("  已保存表格：%s", G_TABLE_PATH)

    # ---- 绘制全局碰撞裕量 G(t) 曲线 ----
    logger.info("步骤 9：绘制全局碰撞裕量 G(t) 曲线")
    fig_path = plot_global_margin(trace, t_star, (i_star, j_star), fine_points)
    logger.info("  已保存图片：%s", fig_path)

    elapsed = time.perf_counter() - t_total
    logger.info("=" * 60)
    logger.info("求解完成，总耗时 %.2f s，终止时刻 t*=%.9f s", elapsed, t_star)
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
