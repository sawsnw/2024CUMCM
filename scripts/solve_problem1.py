"""问题一求解脚本：板凳龙沿等距螺线盘入的位置与速度。

运行方式:
    python scripts/solve_problem1.py

功能:
    1. 按模型求解 t=0..300 s 全部 224 个把手中心的位置与速度；
    2. 读取模板 data/result1.xlsx，将“位置”“速度”工作表结果保存到
       results/tables/result1.xlsx（不修改模板）；
    3. 运行文档第 12 节模型检验并输出误差统计；
    4. 日志写入 results/logs/solve_problem1_YYYY-MM-DD_HHMMSS.log（不覆盖旧文件）。
"""

import os

import openpyxl

from src.config import (
    HEAD_SPEED,
    RESULT1_PATH,
    RESULT1_TEMPLATE_PATH,
    T_END,
    T_START,
    TOTAL_HANDLES,
)
from src.models.spiral_dragon import compute_dragon_at, verify_configuration
from src.utils.logger import setup_logger


def fill_result1(
    template_path: str, output_path: str, results: dict[int, tuple[list, list]]
) -> None:
    """读取模板，将各时刻的位置与速度填入并保存到输出路径。

    模板结构：
        - 位置工作表：第 1 行为时间表头，把手 i 的 x、y 分别在第 2+2i、3+2i 行；
        - 速度工作表：第 1 行为时间表头，把手 i 的速度在第 2+i 行；
        - 时间 t 对应第 t+2 列，数值保留 6 位小数。

    Args:
        template_path: 模板文件路径（data/result1.xlsx，只读不修改）。
        output_path: 结果输出路径（results/tables/result1.xlsx）。
        results: 键为时刻 t，值为 (positions, speeds) 的字典。
    """
    wb = openpyxl.load_workbook(template_path)
    ws_position = wb["位置"]
    ws_speed = wb["速度"]

    for t, (positions, speeds) in results.items():
        col = t + 2
        for i, (x, y) in enumerate(positions):
            ws_position.cell(row=2 + 2 * i, column=col).value = round(x, 6)
            ws_position.cell(row=3 + 2 * i, column=col).value = round(y, 6)
        for i, v in enumerate(speeds):
            ws_speed.cell(row=2 + i, column=col).value = round(v, 6)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


def main() -> None:
    logger = setup_logger(__name__, script_name="solve_problem1")
    logger.info("=" * 60)
    logger.info("问题一：板凳龙沿等距螺线盘入的位置与速度求解")
    logger.info("参数：螺距 p=0.55 m，龙头速度 v=%.1f m/s，初始第 16 圈 (θ=32π)", HEAD_SPEED)
    logger.info("把手总数：%d（龙头前 1 + 龙身 221 + 龙尾前/后 2）", TOTAL_HANDLES)

    # 累计各时刻的检验误差（取全程最大值）
    max_errors = {
        "龙头弧长误差": 0.0,
        "最大弦长误差": 0.0,
        "参数递增": 0.0,
        "龙头速度误差": 0.0,
        "最大速度约束误差": 0.0,
    }
    results: dict[int, tuple[list, list]] = {}

    for t in range(T_START, T_END + 1):
        positions, speeds, thetas, theta_dot, velocities = compute_dragon_at(t)
        results[t] = (positions, speeds)
        ver = verify_configuration(t, thetas, positions, theta_dot, velocities, speeds)
        for key in max_errors:
            max_errors[key] = max(max_errors[key], ver[key])
        if t % 60 == 0:
            head = positions[0]
            logger.info(
                "t=%3d s  龙头前把手: (%.6f, %.6f)  速度 %.6f m/s  参数 θ=%.6f",
                t, head[0], head[1], speeds[0], thetas[0],
            )

    # 输出模型检验汇总
    logger.info("=" * 60)
    logger.info("模型检验汇总（0~300 s 全程最大值）:")
    for key, val in max_errors.items():
        logger.info("  %s: %.3e", key, val)

    # 读取模板、填充并输出结果（模板保持不变）
    logger.info("读取模板：%s", RESULT1_TEMPLATE_PATH)
    logger.info("输出结果：%s", RESULT1_PATH)
    fill_result1(RESULT1_TEMPLATE_PATH, RESULT1_PATH, results)
    logger.info("问题一求解完成，结果已写入 %s", RESULT1_PATH)
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
